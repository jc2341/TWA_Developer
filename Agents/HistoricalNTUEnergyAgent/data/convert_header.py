
'''
This python file does 2 things:
1. Converts the header of the .xlsx file
Below is the header format in the original NTU_Energy_Consumption.xlsx
'NEC_P(KW)'
'NEC_Q(KVAR)'
'CANTEEN_2_P(KW)'

The header format should be converted to below format before our agent can retrieve and map the data
NEC_P_KW
NEC_Q_KVAR
CANTEEN_2_P_KW

2. Convert the time into the format which can be read by our agent
From 1/1/2020  1:00:00 am
To   2020-01-01T01:00:00
'''

'''

'''

import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(filename='NTU_Energy_Consumption.xlsx')
ws = wb.active

# Iterate over the cells in the first row and replace any parentheses with underscores
for cell in ws[1]:
    cell.value = str(cell.value).replace('(', '_').replace(')', '')

# Iterate over the rows starting from the second row
for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    # Convert the date/time string to a datetime object
    datetime_obj = datetime.strptime(str(row[0]), '%Y-%m-%d %H:%M:%S')
    # Format the datetime object to a string in the desired format
    formatted_date_string = datetime_obj.strftime("%Y-%m-%dT%H:%M:%S")
    # Update the cell with the formatted date string
    ws.cell(row=row_number, column=1, value=formatted_date_string)

# Save the updated workbook to a file
wb.save('NTU_Energy_Consumption.xlsx')

